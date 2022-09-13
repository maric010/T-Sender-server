import psycopg2
from telethon.sync import TelegramClient, events
from telethon import functions, types
import socks,time,os,asyncio
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timezone
from telethon.tl.types import InputPhoneContact
from telethon.tl.functions.contacts import ImportContactsRequest
from sys import argv
from PROXY import set_proxy
from DB import  query
script, id = argv

users = []
class USER:
    def __init__(self):
        pass
ACCOUNTS = []
class ACCOUNT:
    def __init__(self):
        self.client=None
        self.message=None


row = query("SELECT SLEEP_TIME,username,text,media FROM mails WHERE id="+id)[0]
SLEEP_TIME = row[0]
username = row[1]
text = row[2]

#187 170 609/maric000/&&755 726 633/maric000/&&
u = None
path = os.getcwd() + "/users/"+username

media = row[3]
if media=="":
    media=None
else:
    media = path+"/"+media
with open(path+'/mails/'+id+'.txt', 'r',encoding="utf-8") as fp:
    u = fp.read().replace(u'\xa0', '').split("&&")
for i in range(len(u)-1):
    su = u[i].split("/")
    new_user = USER()
    try:
        new_user.id = int(su[0])
    except:
        pass
    new_user.username = su[1]
    new_user.phone = su[2]
    users.append(new_user)

query("UPDATE accounts SET status='busy',last_used=now()::timestamp WHERE mail='"+id+"' and owner='"+username+"'")

rows = query("SELECT phone,proxy_ip,proxy_port,proxy_user,proxy_password,proxy_type FROM accounts WHERE mail='"+id+"' and owner='"+username+"'")
GROUPS={}
async def main():
    for row in rows:
        account = ACCOUNT()
        try:
            if row[1]=="empty":
                set_proxy(row[0],username)
                row = query("SELECT phone,proxy_ip,proxy_port,proxy_user,proxy_password,proxy_type FROM accounts WHERE owner='"+username+"' and phone='"+row[0]+"'")[0]
            my_proxy = (socks.HTTP, row[1], int(row[2]), True, row[3], row[4])
            account.proxy = my_proxy
            account.client = TelegramClient(path+"/accounts/+"+row[0], 3442047, "cacb39d73070900d11d07fe14a476394",proxy=account.proxy)

            await account.client.connect()
            ACCOUNTS.append(account)
            if GROUPS.get(my_proxy) is not None:
                GROUPS[my_proxy].append(account)
            else:
                GROUPS[my_proxy] = [account]
            await account.client.disconnect()
        except Exception as e:
            print(e)
    print("Аккаунты загружены. Количество аккаунтов",len(ACCOUNTS))
    print("Количество групп",len(GROUPS))
    step = 0

    active_accounts_count = len(ACCOUNTS)

    a_count = len(users)
    s_count = 0
    f_count = 0

    result_path = path+"/mails/"+id+"_result.xlsx"
    green = PatternFill(start_color='00FF00',end_color='00FF00',fill_type='solid')
    red = PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    wb = Workbook()
    wb.save(result_path)
    def add_result(user,status):
        statik = "Не удалось отправить"
        if status:
            statik = "Успешно"
        wb = load_workbook(result_path)
        ws = wb.active
        ws.append([user.id, user.username,user.phone, statik])
        if status:
            ws['D'+str(ws.max_row)].fill=green
        else:
            ws['D' + str(ws.max_row)].fill = red
        wb.save(result_path)
    query("UPDATE mails SET status='started',start_date=now()::timestamp WHERE id="+id)
    while len(users) > 0:
        for i in range(3):
            if len(users) == 0:
                break
            slp = False
            for group in GROUPS:
                if len(users) == 0:
                    break
                try:
                    if len(GROUPS.get(group))<i+1:
                        print("continue")
                        continue
                    slp=True
                    account = GROUPS.get(group)[i]
                    user = users.pop(0)
                    try:
                        await account.client.connect()
                        try:
                            if account.message == None:
                                account.message = await account.client.send_message("me", text, file=media)
                            if user.username != "":
                                await account.client.send_message(user.username, account.message)
                                print("отправлен с помощи username")
                                s_count += 1
                                add_result(user, True)
                            elif user.phone != "":
                                print(user.phone)
                                try:
                                    entity = await account.client.get_entity(user.phone)
                                    if entity is not None:
                                        await account.client.send_message(entity, account.message)
                                        print("отправлен с помощи phone")
                                        s_count += 1
                                        add_result(user, True)
                                    else:
                                        print(1 / 0)
                                except Exception as e:
                                    print(1, e)
                                    contact = InputPhoneContact(
                                        client_id=0,
                                        phone=user.phone,
                                        first_name=user.name,
                                        last_name=user.surname
                                    )  # For new contacts you should use client_id = 0
                                    result = await account.client(ImportContactsRequest([contact]))
                                    entity = await account.client.get_entity(user.phone)
                                    await account.client.send_message(entity, account.message)
                                    s_count += 1
                                    add_result(user, True)
                            else:
                                await account.client.send_message(int(str(user.id)), account.message)
                                print("отправлен с помощи id")
                                s_count += 1
                                add_result(user, True)
                        except Exception as e:
                            try:
                                if "Too many requests" in str(e):
                                    GROUPS.get(group).remove(account)
                                    active_accounts_count -= 1
                                elif "The file reference has expired" in str(e):
                                    account.message = await account.client.send_message("me", text, file=media)
                                else:
                                    f_count += 1
                                    add_result(user, False)
                            except:
                                pass
                        await account.client.disconnect()
                    except:
                        print(e)

                except Exception as e:
                    print(e)

            query("UPDATE mails SET success_count = '"+str(s_count)+"',failed_count = '"+str(f_count)+"',active_accounts_count = '"+str(active_accounts_count)+"' WHERE id=" + id)
            if len(users) == 0:
                break
            if slp:
                print("sleep")
                asyncio.sleep(SLEEP_TIME)
                print("sleep stop")
    query("UPDATE mails SET status='finished',stop_date=now()::timestamp WHERE id="+id)
    query("UPDATE accounts SET mail='',status='ready',last_used=now()::timestamp WHERE owner='"+username+"' and mail='"+id+"'")

asyncio.run(main())

